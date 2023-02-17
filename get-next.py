import sys
import openpyxl

in_filename = sys.argv[1]

wb = openpyxl.load_workbook(in_filename, data_only=True)
fs = wb.active
fs_count_row = fs.max_row 
fs_count_col = fs.max_column 

full_list = []

for row in range(1, fs_count_row + 1):
    cols = fs_count_col + 1
    column = 1
    while column < cols: 
        cell_color = fs.cell(column=column, row=row)
        
        bgColor = cell_color.fill.bgColor.index
        fgColor = cell_color.fill.fgColor.index
        if (bgColor=='00000000') or (fgColor=='00000000'):
            column += 1 
            continue

        if (fgColor =='FFFFFF00'):
            if str(cell_color.value) != 'None': 
                full_list.append(cell_color.value)
            column += 5
        
        column += 1


full_list.sort()
print('\n'.join(full_list))